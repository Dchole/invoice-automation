from __future__ import annotations
import csv
import io
import random
import shutil
from datetime import date, time, timedelta, datetime
from pathlib import Path
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session as DbSession

from app.database import get_db
from app.models.invoice import Invoice
from app.models.payment import Payment
from app.models.client import Client
from app.models.session import Session as SessionModel
from app.models.reminder import Reminder
from app.services.excel_importer import import_excel
from app.services.csv_importer import import_csv
from app.services.reminder_engine import (
    schedule_reminders,
    schedule_reminders_for_unpaid,
)

router = APIRouter(prefix="/api/import", tags=["import/export"])

UPLOAD_DIR = Path(__file__).resolve().parent.parent.parent.parent / "data"


@router.delete("/reset")
def reset_all_data(db: DbSession = Depends(get_db)):
    """Delete all data from all tables."""
    db.query(Reminder).delete()
    db.query(Payment).delete()
    db.query(SessionModel).delete()
    db.query(Invoice).delete()
    db.query(Client).delete()
    db.commit()
    return {"message": "All data cleared"}


@router.post("/seed")
def seed_data(db: DbSession = Depends(get_db)):
    """Generate demo seed data. Clears existing data first."""
    # --- Clear existing data ---
    db.query(Reminder).delete()
    db.query(Payment).delete()
    db.query(SessionModel).delete()
    db.query(Invoice).delete()
    db.query(Client).delete()
    db.flush()

    random.seed(42)

    # --- Config ---
    NUM_CLIENTS = 100
    NUM_SESSIONS = 240
    MONTHS_SPAN = 4
    INVOICE_CUTOFF = 230
    PAYMENT_RATE = 0.90
    PAYMENT_METHODS = ["e-transfer", "check", "cash", "stripe", "paypal"]

    FIRST_NAMES = [
        "Alice",
        "Bob",
        "Carlos",
        "Diana",
        "Erik",
        "Fiona",
        "George",
        "Hannah",
        "Ivan",
        "Julia",
        "Kevin",
        "Laura",
        "Miguel",
        "Nina",
        "Oscar",
        "Priya",
        "Quinn",
        "Rachel",
        "Sam",
        "Tina",
        "Uma",
        "Victor",
        "Wendy",
        "Xander",
        "Yara",
        "Zach",
        "Amara",
        "Beckett",
        "Celine",
        "Damon",
    ]
    LAST_NAMES = [
        "Anderson",
        "Baker",
        "Chen",
        "Dubois",
        "Ellis",
        "Fernandez",
        "Garcia",
        "Huang",
        "Ibrahim",
        "Jackson",
        "Kim",
        "Larson",
        "Martinez",
        "Nakamura",
        "O'Brien",
        "Patel",
        "Quinn",
        "Rivera",
        "Singh",
        "Thompson",
        "Ueda",
        "Volkov",
        "Williams",
        "Xu",
        "Yamamoto",
        "Zhang",
        "Abadi",
        "Bjornsson",
        "Castillo",
        "Delgado",
    ]
    COMPANIES = [
        "Bright Pixel Studio",
        "Northwind Consulting",
        "Maple & Co",
        "TechForge Inc",
        "Blue Harbor Media",
        "Summit Digital",
        "GreenLeaf Design",
        "Quantum Labs",
        "Sunrise Ventures",
        "Atlas Creative",
        "Pinnacle Solutions",
        "Lakeside Tech",
        "Redwood Strategies",
        "Trailblaze Marketing",
        "Frostbite Games",
        "Coral Reef Apps",
        "Ironclad Security",
        "Velvet Sound Studio",
        "Amber Wave Films",
        "Starlight Events",
        None,
        None,
        None,
        None,
        None,
    ]
    SESSION_DESCRIPTIONS = [
        "Initial project kickoff and scope discussion",
        "UI wireframe review and feedback session",
        "Backend API architecture planning",
        "Database schema design workshop",
        "Sprint planning and task breakdown",
        "Code review and refactoring discussion",
        "Bug triage and prioritization meeting",
        "User testing debrief and action items",
        "Landing page design iteration",
        "Payment integration troubleshooting",
        "SEO audit and recommendations walkthrough",
        "Mobile responsive layout adjustments",
        "Brand guidelines review session",
        "Performance optimization deep-dive",
        "Client onboarding and setup walkthrough",
        "Content strategy brainstorm",
        "Analytics dashboard configuration",
        "Email template design review",
        "Security audit findings discussion",
        "Deployment pipeline setup and testing",
        "Third-party API integration planning",
        "Data migration strategy session",
        "Accessibility compliance review",
        "Social media campaign planning",
        "Quarterly progress review and roadmap update",
        "Invoice workflow automation demo",
        "CRM integration scoping call",
        "Video editing feedback and revisions",
        "Photography selection and layout review",
        "Contract renewal and scope expansion talk",
    ]
    CLIENT_NOTES = [
        "Prefers morning meetings before 10 AM",
        "Key contact for all design approvals",
        "Usually pays within a week, very reliable",
        "Referred by a mutual contact at the tech meetup",
        "Needs detailed time breakdowns on invoices",
        "Prefers email over phone calls",
        "Long-term client, flexible on rates",
        "Startup — budget-conscious but growing fast",
        "Government contract, strict invoicing rules",
        "Always asks for PDF copies of invoices",
        "Responsive and easy to work with",
        "Tends to scope-creep, keep boundaries clear",
        "Pays promptly via e-transfer every time",
        "Multiple projects running in parallel",
        "Seasonal work — busier in Q1 and Q4",
        "Requires bilingual (EN/FR) deliverables",
        "Recently expanded team, more work coming",
        "Non-profit rate applies",
        "Timezone is PST — schedule accordingly",
        "Likes weekly status update emails",
    ]
    INVOICE_NOTES = [
        "Includes all sessions for this billing period",
        "Rush project — premium rate applied",
        "Discounted rate per annual agreement",
        "Final invoice for phase one deliverables",
        "Covers additional revision rounds requested",
        "Travel expenses included in subtotal",
        "Split billing — second half of project",
        "Pro-rated for partial month engagement",
        "Retainer invoice — monthly fixed fee",
        "Overtime hours billed at 1.5x rate",
        "Early payment discount applied (2%)",
        "Adjusted for scope change mid-project",
        "Includes third-party software license cost",
        "Quarterly review and maintenance bundle",
        "Emergency support hours from last week",
    ]
    PAYMENT_NOTES = [
        "Paid on time as usual",
        "Sent confirmation receipt via email",
        "Partial payment — remainder due next month",
        "Paid early, much appreciated",
        "Reference number noted for records",
        "Cleared after a short delay",
        "Payment received with thank-you note",
        "Auto-payment from recurring setup",
        "Paid in full, project complete",
        "Wire transfer confirmed by bank",
        "Check deposited, 3-day hold",
        "Stripe fee deducted from amount",
        "PayPal transaction ID logged",
        "Client paid from a different account",
        "Matched to invoice after manual review",
    ]

    def random_date_in_range(start: date, end: date) -> date:
        delta = (end - start).days
        return start + timedelta(days=random.randint(0, delta))

    # --- Create 100 clients ---
    clients = []
    used_names: set[str] = set()
    for _ in range(NUM_CLIENTS):
        while True:
            first = random.choice(FIRST_NAMES)
            last = random.choice(LAST_NAMES)
            name = f"{first} {last}"
            if name not in used_names:
                used_names.add(name)
                break
        email_last = last.lower().replace("'", "")
        email = f"{first.lower()}.{email_last}@example.com"
        company = random.choice(COMPANIES)
        rate = random.choice([75, 85, 95, 100, 110, 120, 125, 135, 150])
        terms = random.choice([15, 30, 30, 30, 45])
        c = Client(
            name=name,
            email=email,
            company=company,
            currency="CAD",
            default_rate=rate,
            payment_terms=terms,
            notes=random.choice(CLIENT_NOTES),
        )
        db.add(c)
        clients.append(c)
    db.flush()

    # --- Create 240 sessions ---
    month_starts = [
        date(2025, 12, 1),
        date(2026, 1, 1),
        date(2026, 2, 1),
        date(2026, 3, 1),
    ]
    month_ends = [
        date(2025, 12, 31),
        date(2026, 1, 31),
        date(2026, 2, 28),
        date(2026, 3, 31),
    ]
    sessions = []
    for i in range(NUM_SESSIONS):
        month_idx = i % MONTHS_SPAN
        client = random.choice(clients)
        sess_date = random_date_in_range(month_starts[month_idx], month_ends[month_idx])
        duration = random.choice([30, 45, 60, 60, 90, 90, 120])
        rate = float(client.default_rate) if client.default_rate else 100.0
        amount = round(rate * duration / 60, 2)
        hour = random.randint(8, 16)
        minute = random.choice([0, 0, 15, 30, 45])
        start_t = time(hour, minute)
        end_minutes = hour * 60 + minute + duration
        end_t = time(min(end_minutes // 60, 23), end_minutes % 60)
        s = SessionModel(
            client_id=client.id,
            date=sess_date,
            start_time=start_t,
            end_time=end_t,
            duration_minutes=duration,
            hourly_rate=rate,
            amount=amount,
            description=random.choice(SESSION_DESCRIPTIONS),
            status="unbilled",
        )
        db.add(s)
        sessions.append(s)
    db.flush()

    # --- Create invoices ---
    billable = sessions[:INVOICE_CUTOFF]
    client_sessions: dict[int, list] = {}
    for s in billable:
        client_sessions.setdefault(s.client_id, []).append(s)

    invoices = []
    inv_num = 1
    for client_id, sess_list in client_sessions.items():
        sess_list.sort(key=lambda s: s.date)
        chunk_size = random.randint(2, 5)
        for j in range(0, len(sess_list), chunk_size):
            chunk = sess_list[j : j + chunk_size]
            subtotal = sum(float(s.amount) for s in chunk)
            tax_rate = random.choice([0, 5, 13, 13, 15])
            tax_amount = round(subtotal * tax_rate / 100, 2)
            total = round(subtotal + tax_amount, 2)
            issue = max(s.date for s in chunk) + timedelta(days=random.randint(1, 3))
            client = next(c for c in clients if c.id == client_id)
            due = issue + timedelta(days=client.payment_terms)
            inv = Invoice(
                invoice_number=f"INV-{inv_num:04d}",
                client_id=client_id,
                issue_date=issue,
                due_date=due,
                subtotal=subtotal,
                tax_rate=tax_rate,
                tax_amount=tax_amount,
                total=total,
                amount_paid=0,
                currency=client.currency,
                status="sent",
                sent_at=datetime.combine(issue, time(9, 0)),
                notes=random.choice(INVOICE_NOTES),
            )
            db.add(inv)
            db.flush()
            for s in chunk:
                s.status = "billed"
                s.invoice_id = inv.id
            invoices.append(inv)
            inv_num += 1
    db.flush()

    # --- Schedule reminders ---
    for inv in invoices:
        schedule_reminders(db, inv)
    db.flush()

    # --- Pay 90% of invoices ---
    payable = random.sample(invoices, k=int(len(invoices) * PAYMENT_RATE))
    for inv in payable:
        pay_date = inv.issue_date + timedelta(
            days=random.randint(2, inv.due_date.day if inv.due_date.day > 2 else 15)
        )
        pay_date = min(pay_date, inv.due_date + timedelta(days=5))
        method = random.choice(PAYMENT_METHODS)
        ref_prefix = {
            "e-transfer": "ET",
            "check": "CHK",
            "cash": "CASH",
            "stripe": "STR",
            "paypal": "PP",
        }
        ref = f"{ref_prefix[method]}-{random.randint(100000, 999999)}"
        p = Payment(
            invoice_id=inv.id,
            amount=float(inv.total),
            payment_date=pay_date,
            payment_method=method,
            reference=ref,
            notes=random.choice(PAYMENT_NOTES),
        )
        db.add(p)
        inv.amount_paid = float(inv.total)
        inv.status = "paid"
        inv.paid_at = datetime.combine(pay_date, time(14, 0))

    today = date.today()
    for inv in invoices:
        if inv.status != "paid" and inv.due_date < today:
            inv.status = "overdue"

    # Skip reminders for paid invoices
    for inv in invoices:
        if inv.status == "paid":
            for r in db.query(Reminder).filter(Reminder.invoice_id == inv.id).all():
                r.status = "skipped"

    db.commit()

    total_invoices = len(invoices)
    paid_count = sum(1 for inv in invoices if inv.status == "paid")
    total_reminders = db.query(Reminder).count()
    pending_reminders = db.query(Reminder).filter(Reminder.status == "pending").count()

    return {
        "clients": len(clients),
        "sessions": len(sessions),
        "invoices": total_invoices,
        "invoices_paid": paid_count,
        "invoices_unpaid": total_invoices - paid_count,
        "payments": len(payable),
        "reminders": total_reminders,
        "reminders_pending": pending_reminders,
    }


@router.post("/excel")
def upload_excel(file: UploadFile = File(...), db: DbSession = Depends(get_db)):
    safe_filename = Path(file.filename).name
    if not safe_filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx or .xls files are supported")

    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    dest = UPLOAD_DIR / f"import_{safe_filename}"
    try:
        with open(dest, "wb") as f:
            shutil.copyfileobj(file.file, f)
        result = import_excel(db, str(dest))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Failed to process file: {e}")
    reminders_created = schedule_reminders_for_unpaid(db)
    db.commit()
    return {
        "clients_created": result.clients_created,
        "sessions_created": result.sessions_created,
        "invoices_created": result.invoices_created,
        "payments_created": result.payments_created,
        "reminders_created": reminders_created,
        "errors": result.errors,
        "warnings": result.warnings,
    }


@router.post("/csv")
def upload_csv(file: UploadFile = File(...), db: DbSession = Depends(get_db)):
    filename = (file.filename or "").lower()
    if not filename.endswith(".csv"):
        raise HTTPException(400, "Only .csv files are supported")

    try:
        content = file.file.read().decode("utf-8-sig")
    except UnicodeDecodeError:
        raise HTTPException(400, "CSV file must be UTF-8 encoded")
    result = import_csv(db, content)
    reminders_created = schedule_reminders_for_unpaid(db)
    db.commit()
    return {
        "clients_created": result.clients_created,
        "sessions_created": result.sessions_created,
        "reminders_created": reminders_created,
        "errors": result.errors,
        "warnings": result.warnings,
    }


@router.get("/export/backup")
def export_backup():
    db_path = UPLOAD_DIR / "invoices.db"
    if not db_path.exists():
        raise HTTPException(404, "Database not found")
    return FileResponse(
        str(db_path),
        filename="invoices_backup.db",
        media_type="application/octet-stream",
    )


@router.get("/export/invoices-csv")
def export_invoices_csv(db: DbSession = Depends(get_db)):
    """Export all invoices as CSV for accounting software."""
    invoices = db.query(Invoice).order_by(Invoice.issue_date.desc()).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "Invoice Number",
            "Client",
            "Client Email",
            "Issue Date",
            "Due Date",
            "Subtotal",
            "Tax Rate %",
            "Tax Amount",
            "Total",
            "Amount Paid",
            "Balance",
            "Currency",
            "Status",
            "Sent Date",
            "Paid Date",
        ]
    )
    for inv in invoices:
        client = db.get(Client, inv.client_id)
        balance = float(inv.total) - float(inv.amount_paid)
        writer.writerow(
            [
                inv.invoice_number,
                client.name if client else "",
                client.email if client and client.email else "",
                str(inv.issue_date),
                str(inv.due_date),
                f"{float(inv.subtotal):.2f}",
                f"{float(inv.tax_rate):.2f}",
                f"{float(inv.tax_amount):.2f}",
                f"{float(inv.total):.2f}",
                f"{float(inv.amount_paid):.2f}",
                f"{balance:.2f}",
                inv.currency,
                inv.status,
                str(inv.sent_at or ""),
                str(inv.paid_at or ""),
            ]
        )
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=invoices_export.csv"},
    )


@router.get("/export/payments-csv")
def export_payments_csv(db: DbSession = Depends(get_db)):
    """Export all payments as CSV for accounting software."""
    payments = db.query(Payment).order_by(Payment.payment_date.desc()).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "Date",
            "Invoice Number",
            "Client",
            "Client Email",
            "Amount",
            "Payment Method",
            "Reference",
            "Notes",
        ]
    )
    for p in payments:
        inv = db.get(Invoice, p.invoice_id)
        client = db.get(Client, inv.client_id) if inv else None
        writer.writerow(
            [
                str(p.payment_date),
                inv.invoice_number if inv else "",
                client.name if client else "",
                client.email if client and client.email else "",
                f"{float(p.amount):.2f}",
                p.payment_method or "",
                p.reference or "",
                p.notes or "",
            ]
        )
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=payments_export.csv"},
    )


@router.get("/export/invoices-excel")
def export_invoices_excel(db: DbSession = Depends(get_db)):
    """Export all invoices as Excel (.xlsx) for accounting software."""
    invoices = db.query(Invoice).order_by(Invoice.issue_date.desc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    ws.append(
        [
            "Invoice Number",
            "Client",
            "Client Email",
            "Issue Date",
            "Due Date",
            "Subtotal",
            "Tax Rate %",
            "Tax Amount",
            "Total",
            "Amount Paid",
            "Balance",
            "Currency",
            "Status",
            "Sent Date",
            "Paid Date",
        ]
    )
    for inv in invoices:
        client = db.get(Client, inv.client_id)
        balance = float(inv.total) - float(inv.amount_paid)
        ws.append(
            [
                inv.invoice_number,
                client.name if client else "",
                client.email if client and client.email else "",
                str(inv.issue_date),
                str(inv.due_date),
                float(inv.subtotal),
                float(inv.tax_rate),
                float(inv.tax_amount),
                float(inv.total),
                float(inv.amount_paid),
                balance,
                inv.currency,
                inv.status,
                str(inv.sent_at or ""),
                str(inv.paid_at or ""),
            ]
        )
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=invoices_export.xlsx"},
    )


@router.get("/export/payments-excel")
def export_payments_excel(db: DbSession = Depends(get_db)):
    """Export all payments as Excel (.xlsx) for accounting software."""
    payments = db.query(Payment).order_by(Payment.payment_date.desc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    ws.append(
        [
            "Date",
            "Invoice Number",
            "Client",
            "Client Email",
            "Amount",
            "Payment Method",
            "Reference",
            "Notes",
        ]
    )
    for p in payments:
        inv = db.get(Invoice, p.invoice_id)
        client = db.get(Client, inv.client_id) if inv else None
        ws.append(
            [
                str(p.payment_date),
                inv.invoice_number if inv else "",
                client.name if client else "",
                client.email if client and client.email else "",
                float(p.amount),
                p.payment_method or "",
                p.reference or "",
                p.notes or "",
            ]
        )
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=payments_export.xlsx"},
    )


@router.get("/export/excel")
def export_excel(db: DbSession = Depends(get_db)):
    """Export all data as a single .xlsx workbook that can be re-imported."""
    wb = Workbook()

    # --- Client Summary sheet ---
    ws_clients = wb.active
    ws_clients.title = "Client Summary"
    ws_clients.append(["Client Name", "Email", "Currency", "Rate", "Payment Terms"])
    clients = db.query(Client).order_by(Client.name).all()
    client_map: dict[int, Client] = {}
    for c in clients:
        client_map[c.id] = c
        ws_clients.append(
            [
                c.name,
                c.email or "",
                c.currency,
                float(c.default_rate) if c.default_rate else "",
                c.payment_terms,
            ]
        )

    # --- Invoices sheet ---
    ws_inv = wb.create_sheet("Invoices")
    ws_inv.append(
        [
            "Invoice Number",
            "Client",
            "Client Email",
            "Issue Date",
            "Due Date",
            "Subtotal",
            "Tax Rate %",
            "Tax Amount",
            "Total",
            "Amount Paid",
            "Balance",
            "Currency",
            "Status",
            "Sent Date",
            "Paid Date",
        ]
    )
    invoices = db.query(Invoice).order_by(Invoice.issue_date.desc()).all()
    for inv in invoices:
        client = client_map.get(inv.client_id)
        balance = float(inv.total) - float(inv.amount_paid)
        ws_inv.append(
            [
                inv.invoice_number,
                client.name if client else "",
                client.email if client and client.email else "",
                inv.issue_date,
                inv.due_date,
                float(inv.subtotal),
                float(inv.tax_rate),
                float(inv.tax_amount),
                float(inv.total),
                float(inv.amount_paid),
                round(balance, 2),
                inv.currency,
                inv.status,
                str(inv.sent_at.date()) if inv.sent_at else "",
                str(inv.paid_at.date()) if inv.paid_at else "",
            ]
        )

    # --- Payments sheet ---
    ws_pay = wb.create_sheet("Payments")
    ws_pay.append(
        [
            "Date",
            "Invoice Number",
            "Client",
            "Client Email",
            "Amount",
            "Payment Method",
            "Reference",
            "Notes",
        ]
    )
    payments = db.query(Payment).order_by(Payment.payment_date.desc()).all()
    for p in payments:
        inv = db.get(Invoice, p.invoice_id)
        client = client_map.get(inv.client_id) if inv else None
        ws_pay.append(
            [
                p.payment_date,
                inv.invoice_number if inv else "",
                client.name if client else "",
                client.email if client and client.email else "",
                float(p.amount),
                p.payment_method or "",
                p.reference or "",
                p.notes or "",
            ]
        )

    # --- Sessions sheet ---
    ws_sess = wb.create_sheet("Sessions")
    ws_sess.append(
        [
            "Client",
            "Client Email",
            "Date",
            "Duration",
            "Rate",
            "Amount",
            "Description",
            "Status",
        ]
    )
    sessions = db.query(SessionModel).order_by(SessionModel.date.desc()).all()
    for s in sessions:
        client = client_map.get(s.client_id)
        ws_sess.append(
            [
                client.name if client else "",
                client.email if client and client.email else "",
                s.date,
                s.duration_minutes,
                float(s.hourly_rate) if s.hourly_rate else "",
                float(s.amount) if s.amount else "",
                s.description or "",
                s.status,
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = date.today().isoformat()
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=invoice_automation_export_{today}.xlsx"
        },
    )


@router.get("/export/csv")
def export_csv(db: DbSession = Depends(get_db)):
    """Export all data as a single CSV (multiple sections separated by headers)."""
    output = io.StringIO()
    writer = csv.writer(output)

    # --- Clients ---
    clients = db.query(Client).order_by(Client.name).all()
    client_map: dict[int, Client] = {}
    writer.writerow(["## Clients"])
    writer.writerow(["Client Name", "Email", "Currency", "Rate", "Payment Terms"])
    for c in clients:
        client_map[c.id] = c
        writer.writerow(
            [
                c.name,
                c.email or "",
                c.currency,
                f"{float(c.default_rate):.2f}" if c.default_rate else "",
                c.payment_terms,
            ]
        )

    # --- Invoices ---
    writer.writerow([])
    writer.writerow(["## Invoices"])
    writer.writerow(
        [
            "Invoice Number",
            "Client",
            "Client Email",
            "Issue Date",
            "Due Date",
            "Subtotal",
            "Tax Rate %",
            "Tax Amount",
            "Total",
            "Amount Paid",
            "Balance",
            "Currency",
            "Status",
            "Sent Date",
            "Paid Date",
        ]
    )
    invoices = db.query(Invoice).order_by(Invoice.issue_date.desc()).all()
    for inv in invoices:
        client = client_map.get(inv.client_id)
        balance = float(inv.total) - float(inv.amount_paid)
        writer.writerow(
            [
                inv.invoice_number,
                client.name if client else "",
                client.email if client and client.email else "",
                str(inv.issue_date),
                str(inv.due_date),
                f"{float(inv.subtotal):.2f}",
                f"{float(inv.tax_rate):.2f}",
                f"{float(inv.tax_amount):.2f}",
                f"{float(inv.total):.2f}",
                f"{float(inv.amount_paid):.2f}",
                f"{balance:.2f}",
                inv.currency,
                inv.status,
                str(inv.sent_at or ""),
                str(inv.paid_at or ""),
            ]
        )

    # --- Payments ---
    writer.writerow([])
    writer.writerow(["## Payments"])
    writer.writerow(
        [
            "Date",
            "Invoice Number",
            "Client",
            "Client Email",
            "Amount",
            "Payment Method",
            "Reference",
            "Notes",
        ]
    )
    payments = db.query(Payment).order_by(Payment.payment_date.desc()).all()
    for p in payments:
        inv = db.get(Invoice, p.invoice_id)
        client = client_map.get(inv.client_id) if inv else None
        writer.writerow(
            [
                str(p.payment_date),
                inv.invoice_number if inv else "",
                client.name if client else "",
                client.email if client and client.email else "",
                f"{float(p.amount):.2f}",
                p.payment_method or "",
                p.reference or "",
                p.notes or "",
            ]
        )

    # --- Sessions ---
    writer.writerow([])
    writer.writerow(["## Sessions"])
    writer.writerow(
        [
            "Client",
            "Client Email",
            "Date",
            "Duration",
            "Rate",
            "Amount",
            "Description",
            "Status",
        ]
    )
    sessions = db.query(SessionModel).order_by(SessionModel.date.desc()).all()
    for s in sessions:
        client = client_map.get(s.client_id)
        writer.writerow(
            [
                client.name if client else "",
                client.email if client and client.email else "",
                str(s.date),
                s.duration_minutes,
                f"{float(s.hourly_rate):.2f}" if s.hourly_rate else "",
                f"{float(s.amount):.2f}" if s.amount else "",
                s.description or "",
                s.status,
            ]
        )

    output.seek(0)
    today = date.today().isoformat()
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=invoice_automation_export_{today}.csv"
        },
    )
