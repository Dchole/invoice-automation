from __future__ import annotations
import re
from datetime import datetime, date, time
from typing import Optional
from pathlib import Path
from openpyxl import load_workbook
from sqlalchemy.orm import Session as DbSession

from app.models.client import Client
from app.models.session import Session
from app.models.invoice import Invoice
from app.models.payment import Payment


def _clean_rate(val) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("$", "").replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def _parse_duration(val, start_time=None, end_time=None) -> Optional[int]:
    if val is not None:
        try:
            return int(float(val))
        except (ValueError, TypeError):
            pass
    if start_time and end_time:
        if isinstance(start_time, time) and isinstance(end_time, time):
            start_dt = datetime.combine(datetime.min, start_time)
            end_dt = datetime.combine(datetime.min, end_time)
            diff = (end_dt - start_dt).total_seconds() / 60
            if diff > 0:
                return int(diff)
    return None


def _parse_date(val) -> Optional[date]:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%B %d, %Y", "%b %d, %Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _parse_time(val) -> Optional[time]:
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, str):
        for fmt in ("%H:%M", "%I:%M %p", "%H:%M:%S"):
            try:
                return datetime.strptime(val.strip(), fmt).time()
            except ValueError:
                continue
    return None


# Map common sheet names to session statuses
STATUS_MAP = {
    "paid": "paid",
    "payment received": "paid",
    "invoiced": "invoiced",
    "invoice sent": "invoiced",
    "sent": "invoiced",
    "unbilled": "unbilled",
    "not invoiced": "unbilled",
    "pending": "unbilled",
    "issues": "unbilled",
    "clients with issues": "unbilled",
}


def _guess_status(sheet_name: str) -> str:
    lower = sheet_name.lower().strip()
    for key, status in STATUS_MAP.items():
        if key in lower:
            return status
    return "unbilled"


class ImportResult:
    def __init__(self):
        self.clients_created: int = 0
        self.sessions_created: int = 0
        self.invoices_created: int = 0
        self.payments_created: int = 0
        self.errors: list[str] = []
        self.warnings: list[str] = []


def _import_client_summary(db: DbSession, wb, client_cache: dict, result: ImportResult):
    """Pre-pass: read Client Summary sheet to populate client details (email, currency, rate)."""
    summary_names = ["client summary", "clients", "client list", "summary"]
    for sheet_name in wb.sheetnames:
        if sheet_name.lower().strip() in summary_names:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(min_row=1, values_only=False))
            if not rows:
                continue

            headers = [str(c.value or "").strip().lower() for c in rows[0]]
            col = {}
            for i, h in enumerate(headers):
                if "client" in h or "name" in h:
                    col.setdefault("name", i)
                elif "email" in h:
                    col.setdefault("email", i)
                elif "currency" in h:
                    col.setdefault("currency", i)
                elif "rate" in h:
                    col.setdefault("rate", i)
                elif "terms" in h or "payment" in h:
                    col.setdefault("terms", i)

            if "name" not in col:
                continue

            for row in rows[1:]:
                vals = [c.value for c in row]
                name = str(vals[col["name"]] or "").strip()
                if not name:
                    continue

                key = name.lower()
                if key not in client_cache:
                    rate = _clean_rate(vals[col["rate"]]) if "rate" in col else None
                    client = Client(name=name, default_rate=rate)
                    db.add(client)
                    db.flush()
                    client_cache[key] = client
                    result.clients_created += 1

                client = client_cache[key]
                if "email" in col and vals[col["email"]]:
                    email = str(vals[col["email"]]).strip()
                    if email and "@" in email:
                        client.email = email
                if "currency" in col and vals[col["currency"]]:
                    currency = str(vals[col["currency"]]).strip().upper()
                    if currency in ("CAD", "USD", "EUR", "GBP"):
                        client.currency = currency
                if "rate" in col and vals[col["rate"]]:
                    rate = _clean_rate(vals[col["rate"]])
                    if rate:
                        client.default_rate = rate
                if "terms" in col and vals[col["terms"]]:
                    try:
                        client.payment_terms = int(float(vals[col["terms"]]))
                    except (ValueError, TypeError):
                        pass

            db.flush()
            break  # Only process first matching summary sheet


def _import_invoices(db: DbSession, sheet_name: str, rows: list, headers: list[str],
                     client_cache: dict, result: ImportResult):
    """Import invoice rows, creating Invoice records matched to clients by name."""
    col = {}
    for i, h in enumerate(headers):
        if "invoice" in h and "number" in h:
            col.setdefault("invoice_number", i)
        elif "client" in h or h == "name":
            col.setdefault("client", i)
        elif "issue" in h and "date" in h:
            col.setdefault("issue_date", i)
        elif "due" in h and "date" in h:
            col.setdefault("due_date", i)
        elif "subtotal" in h:
            col.setdefault("subtotal", i)
        elif "tax" in h and ("%" in h or "rate" in h):
            col.setdefault("tax_rate", i)
        elif "tax" in h and "amount" in h:
            col.setdefault("tax_amount", i)
        elif h in ("total",):
            col.setdefault("total", i)
        elif "amount paid" in h or "paid" in h and "date" not in h:
            col.setdefault("amount_paid", i)
        elif "balance" in h:
            col.setdefault("balance", i)
        elif "currency" in h:
            col.setdefault("currency", i)
        elif "status" in h:
            col.setdefault("status", i)
        elif "sent" in h:
            col.setdefault("sent_at", i)
        elif h == "paid date":
            col.setdefault("paid_at", i)

    if "invoice_number" not in col:
        result.warnings.append(f"Sheet '{sheet_name}': no invoice number column, skipping")
        return

    for row_idx, row in enumerate(rows[1:], start=2):
        vals = [c.value for c in row]
        inv_num = str(vals[col["invoice_number"]] or "").strip()
        if not inv_num:
            continue

        # Skip if invoice already exists
        existing = db.query(Invoice).filter(Invoice.invoice_number == inv_num).first()
        if existing:
            result.warnings.append(f"Sheet '{sheet_name}' row {row_idx}: invoice '{inv_num}' already exists, skipping")
            continue

        # Resolve client
        client = None
        if "client" in col:
            client_name = str(vals[col["client"]] or "").strip()
            if client_name:
                key = client_name.lower()
                if key not in client_cache:
                    c = Client(name=client_name)
                    db.add(c)
                    db.flush()
                    client_cache[key] = c
                    result.clients_created += 1
                client = client_cache[key]

        if not client:
            result.errors.append(f"Sheet '{sheet_name}' row {row_idx}: no client for invoice '{inv_num}'")
            continue

        issue_date = _parse_date(vals[col["issue_date"]]) if "issue_date" in col else None
        due_date = _parse_date(vals[col["due_date"]]) if "due_date" in col else None
        if not issue_date or not due_date:
            result.errors.append(f"Sheet '{sheet_name}' row {row_idx}: missing issue/due date")
            continue

        subtotal = _clean_rate(vals[col["subtotal"]]) if "subtotal" in col else 0
        tax_rate = _clean_rate(vals[col["tax_rate"]]) if "tax_rate" in col else 0
        tax_amount = _clean_rate(vals[col["tax_amount"]]) if "tax_amount" in col else 0
        total = _clean_rate(vals[col["total"]]) if "total" in col else (subtotal or 0) + (tax_amount or 0)
        amount_paid = _clean_rate(vals[col["amount_paid"]]) if "amount_paid" in col else 0
        currency_val = str(vals[col["currency"]]).strip().upper() if "currency" in col and vals[col["currency"]] else client.currency
        status_val = str(vals[col["status"]]).strip().lower() if "status" in col and vals[col["status"]] else "draft"

        sent_at = None
        if "sent_at" in col and vals[col["sent_at"]]:
            d = _parse_date(vals[col["sent_at"]])
            if d:
                from datetime import datetime as dt
                sent_at = dt.combine(d, dt.min.time())

        paid_at = None
        if "paid_at" in col and vals[col["paid_at"]]:
            d = _parse_date(vals[col["paid_at"]])
            if d:
                from datetime import datetime as dt
                paid_at = dt.combine(d, dt.min.time())

        invoice = Invoice(
            invoice_number=inv_num,
            client_id=client.id,
            issue_date=issue_date,
            due_date=due_date,
            subtotal=subtotal or 0,
            tax_rate=tax_rate or 0,
            tax_amount=tax_amount or 0,
            total=total or 0,
            amount_paid=amount_paid or 0,
            currency=currency_val,
            status=status_val,
            sent_at=sent_at,
            paid_at=paid_at,
        )
        db.add(invoice)
        db.flush()
        result.invoices_created += 1


def _import_payments(db: DbSession, sheet_name: str, rows: list, headers: list[str], result: ImportResult):
    """Import payment rows by matching invoice numbers to existing invoices."""
    col = {}
    for i, h in enumerate(headers):
        if "invoice" in h and "number" in h:
            col.setdefault("invoice_number", i)
        elif "date" in h:
            col.setdefault("date", i)
        elif "amount" in h:
            col.setdefault("amount", i)
        elif "method" in h:
            col.setdefault("method", i)
        elif "reference" in h:
            col.setdefault("reference", i)
        elif "note" in h:
            col.setdefault("notes", i)

    if "invoice_number" not in col or "amount" not in col:
        result.warnings.append(f"Sheet '{sheet_name}': missing invoice number or amount column, skipping")
        return

    for row_idx, row in enumerate(rows[1:], start=2):
        vals = [c.value for c in row]
        inv_num = str(vals[col["invoice_number"]] or "").strip()
        if not inv_num:
            continue

        invoice = db.query(Invoice).filter(Invoice.invoice_number == inv_num).first()
        if not invoice:
            result.errors.append(f"Sheet '{sheet_name}' row {row_idx}: invoice '{inv_num}' not found")
            continue

        amount = _clean_rate(vals[col["amount"]])
        if not amount or amount <= 0:
            result.errors.append(f"Sheet '{sheet_name}' row {row_idx}: invalid amount")
            continue

        payment_date = _parse_date(vals[col["date"]]) if "date" in col else None
        if not payment_date:
            result.errors.append(f"Sheet '{sheet_name}' row {row_idx}: invalid or missing date")
            continue

        method = str(vals[col["method"]]).strip() if "method" in col and vals[col["method"]] else None
        reference = str(vals[col["reference"]]).strip() if "reference" in col and vals[col["reference"]] else None
        notes = str(vals[col["notes"]]).strip() if "notes" in col and vals[col["notes"]] else None

        # Check for duplicate: same invoice, date, and amount
        existing = db.query(Payment).filter(
            Payment.invoice_id == invoice.id,
            Payment.payment_date == payment_date,
            Payment.amount == amount,
        ).first()
        if existing:
            result.warnings.append(f"Sheet '{sheet_name}' row {row_idx}: duplicate payment skipped")
            continue

        payment = Payment(
            invoice_id=invoice.id,
            amount=amount,
            payment_date=payment_date,
            payment_method=method,
            reference=reference,
            notes=notes,
        )
        db.add(payment)
        result.payments_created += 1


def import_excel(db: DbSession, file_path: str) -> ImportResult:
    result = ImportResult()
    wb = load_workbook(file_path, data_only=True)
    client_cache: dict[str, Client] = {}

    for existing in db.query(Client).all():
        client_cache[existing.name.lower()] = existing

    # Pre-pass: read client summary sheet for emails, currency, rates
    _import_client_summary(db, wb, client_cache, result)

    # Categorise sheets
    invoice_sheets = []
    payment_sheets = []
    session_sheets = []
    client_summary_names = {"client summary", "clients", "client list", "summary"}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, values_only=False))
        if not rows:
            continue

        headers = [str(c.value or "").strip().lower() for c in rows[0]]
        header_set = set(headers)
        header_joined = " ".join(headers)

        if sheet_name.lower().strip() in client_summary_names:
            continue  # Already processed in pre-pass

        payment_markers = {"payment method", "reference", "invoice number"}
        invoice_markers = {"invoice number", "due date", "tax rate", "tax rate %", "tax amount", "subtotal"}
        is_payment_sheet = len(payment_markers & header_set) >= 2 and "duration" not in header_joined
        is_invoice_sheet = len(invoice_markers & header_set) >= 3 and "duration" not in header_joined

        if is_invoice_sheet and not is_payment_sheet:
            invoice_sheets.append((sheet_name, rows, headers))
        elif is_payment_sheet:
            payment_sheets.append((sheet_name, rows, headers))
        else:
            session_sheets.append((sheet_name, rows, headers))

    # Pass 1: invoices (so payment references resolve)
    for sheet_name, rows, headers in invoice_sheets:
        _import_invoices(db, sheet_name, rows, headers, client_cache, result)
    db.flush()

    # Pass 2: payments
    for sheet_name, rows, headers in payment_sheets:
        _import_payments(db, sheet_name, rows, headers, result)

    # Pass 3: sessions
    for sheet_name, rows, headers in session_sheets:
        status = _guess_status(sheet_name)

        col_map = {}
        for i, h in enumerate(headers):  # noqa: E501
            if "client" in h or "name" in h:
                col_map.setdefault("client", i)
            elif "date" in h and "pay" not in h:
                col_map.setdefault("date", i)
            elif "start" in h:
                col_map.setdefault("start", i)
            elif "end" in h:
                col_map.setdefault("end", i)
            elif "duration" in h or "minutes" in h or "mins" in h:
                col_map.setdefault("duration", i)
            elif "rate" in h:
                col_map.setdefault("rate", i)
            elif "amount" in h or "total" in h or "fee" in h:
                col_map.setdefault("amount", i)
            elif "desc" in h or "note" in h or "service" in h or "topic" in h:
                col_map.setdefault("description", i)

        if "client" not in col_map:
            result.warnings.append(f"Sheet '{sheet_name}': no client column found, skipping")
            continue

        for row_idx, row in enumerate(rows[1:], start=2):
            vals = [c.value for c in row]
            client_name = str(vals[col_map["client"]] or "").strip()
            if not client_name:
                continue

            if client_name.lower() not in client_cache:
                rate = _clean_rate(vals[col_map["rate"]]) if "rate" in col_map else None
                client = Client(name=client_name, default_rate=rate)
                db.add(client)
                db.flush()
                client_cache[client_name.lower()] = client
                result.clients_created += 1

            client = client_cache[client_name.lower()]

            session_date = _parse_date(vals[col_map["date"]]) if "date" in col_map else None
            if not session_date:
                result.errors.append(f"Sheet '{sheet_name}' row {row_idx}: invalid or missing date")
                continue

            start = _parse_time(vals[col_map["start"]]) if "start" in col_map else None
            end = _parse_time(vals[col_map["end"]]) if "end" in col_map else None
            duration = _parse_duration(
                vals[col_map["duration"]] if "duration" in col_map else None,
                start, end,
            )
            if duration is None or duration <= 0:
                result.errors.append(f"Sheet '{sheet_name}' row {row_idx}: cannot determine duration")
                continue

            rate = _clean_rate(vals[col_map["rate"]]) if "rate" in col_map else None
            if rate is None:
                rate = float(client.default_rate) if client.default_rate else None
            if rate is None:
                result.errors.append(f"Sheet '{sheet_name}' row {row_idx}: no rate found")
                continue

            amount = round(duration / 60.0 * rate, 2)
            desc = str(vals[col_map["description"]]) if "description" in col_map and vals[col_map["description"]] else None

            session = Session(
                client_id=client.id,
                date=session_date,
                start_time=start,
                end_time=end,
                duration_minutes=duration,
                hourly_rate=rate,
                amount=amount,
                description=desc,
                status=status,
            )
            db.add(session)
            result.sessions_created += 1

    db.commit()
    return result
